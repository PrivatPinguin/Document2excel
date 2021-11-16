import openpyxl

#find first emtpy row in the sheet and returns the row number
def find_empty_row(sheet):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value is None:
            return row
    return sheet.max_row + 1

#use find_empty_row to write data to the sheet
def write_to_excel(item):
    filename = 'unbenannt 1.xlsx'
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    row = find_empty_row(sheet)
    # for item in data:
    sheet.cell(row, 1).value = item[0]
    sheet.cell(row, 2).value = item[1]
    sheet.cell(row, 3).value = item[2]
    sheet.cell(row, 4).value = item[3]
    sheet.cell(row, 5).value = item[4]
        # row += 1
    wb.save(filename)


# # test write_to_excel with "unbenannt 1.xlsx"
# # use some random string to append to the filename
# # to avoid overwriting existing files
# filename = 'unbenannt 1.xlsx'
# data = [
#     ['Max', 'Mustermann', '01.01.2000'],
#     ['', 'Müller', '01.01.2000'],
#     ['', 'Meier', '01.01.2000'],
#     ['', 'Müller', '01.01.2000'],
# ]
# write_to_excel(data, filename)
